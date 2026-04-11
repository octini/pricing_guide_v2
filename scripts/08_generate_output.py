#!/usr/bin/env python3
"""Phase 8: Generate Excel and CSV output"""

import sys
import csv
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

INPUT_CSV = Path('data/processed/items_validated.csv')
OUTPUT_XLSX = Path('output/pricing_guide.xlsx')
OUTPUT_CSV = Path('output/pricing_guide.csv')

RARITY_COLORS = {
    'mundane': 'F5F5F5',
    'common': 'FFFFFF',
    'uncommon': '1EFF00',
    'rare': '0070DD',
    'very_rare': 'A335EE',
    'legendary': 'FF8000',
    'artifact': 'E6CC80',
    'unknown_magic': 'DDDDDD',
    'unknown': 'EEEEEE',
    'varies': 'BBBBBB',
}

RARITY_TEXT_COLORS = {
    'uncommon': '000000',
    'rare': 'FFFFFF',
    'very_rare': 'FFFFFF',
    'legendary': '000000',
    'artifact': '000000',
}

def format_price(price_gp):
    if price_gp < 1:
        return f"{int(price_gp * 100)} cp"
    elif price_gp < 10:
        return f"{price_gp:.1f} gp"
    else:
        return f"{int(price_gp):,} gp"

def determine_price_source_label(row):
    if row.get('price_source') == 'official':
        return 'Official'
    elif row.get('price_confidence') == 'multi':
        return f"Amalgamated ({row.get('price_sources', '')})"
    elif row.get('price_confidence') == 'solo':
        return f"Single source ({row.get('price_sources', '')})"
    else:
        return 'Algorithm'

def main():
    df = pd.read_csv(INPUT_CSV)
    print(f'Loaded {len(df)} items')
    
    # Exclude generic variants (items with 'items' field in raw JSON)
    # These are placeholder items like "Horn of Valhalla" that have specific variants
    if 'is_generic_variant' in df.columns:
        before = len(df)
        df = df[~df['is_generic_variant'].fillna(False)]
        print(f'Excluded {before - len(df)} generic variants')

    output_rows = []
    for _, row in df.iterrows():
        price = row.get('final_price', row.get('rule_price', 0))
        price_low = row.get('price_low', 0)
        price_high = row.get('price_high', 0)
        confidence = row.get('confidence', '')
        # Build notes - flag items without reference sources
        notes = ''
        if row.get('is_outlier', False):
            notes = '⚠️'
        if not row.get('has_reference_source', True):
            notes = '🤖' if notes else '🤖'  # Robot emoji for algorithmic items
        
        output_rows.append({
            'Name': row['name'],
            'Source': row['source'],
            'Type': row.get('type', ''),
            'Rarity': row['rarity'].replace('_', ' ').title(),
            'Attunement': row.get('req_attune', 'none').replace('none', 'No'),
            'Price (gp)': round(float(price), 2) if pd.notna(price) else 0,
            'Price Formatted': format_price(float(price)) if pd.notna(price) else '0 gp',
            'Price Low': format_price(float(price_low)) if pd.notna(price_low) and price_low > 0 else '',
            'Price High': format_price(float(price_high)) if pd.notna(price_high) and price_high > 0 else '',
            'Confidence': confidence,
            'Price Source': determine_price_source_label(row.to_dict()),
            'URL': row.get('url', ''),
            'Is Outlier': row.get('is_outlier', False),
            'Has Reference': row.get('has_reference_source', True),
        })

    out_df = pd.DataFrame(output_rows)

    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    # For CSV, include additional columns beyond the main view
    csv_df = out_df.drop(columns=['URL', 'Is Outlier'])
    csv_df.to_csv(OUTPUT_CSV, index=False, quoting=csv.QUOTE_ALL)
    print(f'Saved CSV to {OUTPUT_CSV}')

    # Create multi-sheet Excel workbook
    wb = openpyxl.Workbook()
    
    # Sheet 1: Main Pricing Guide
    ws_main = wb.active
    ws_main.title = 'Pricing Guide'

    headers = ['Name', 'Source', 'Type', 'Rarity', 'Attunement', 'Price (gp)', 'Price Low', 'Price High', 'Confidence', 'Price Source', 'Notes']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=col_idx, value=header)
        cell.fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')

    ws_main.freeze_panes = 'A2'

    for row_idx, row in enumerate(output_rows, 2):
        rarity_key = row['Rarity'].lower().replace(' ', '_')
        bg_color = RARITY_COLORS.get(rarity_key, 'FFFFFF')
        text_color = RARITY_TEXT_COLORS.get(rarity_key, '000000')

        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        font_normal = Font(color=text_color, size=10)
        font_link = Font(color='0563C1', underline='single', size=10)

        name_cell = ws_main.cell(row=row_idx, column=1, value=row['Name'])
        if row['URL']:
            name_cell.hyperlink = row['URL']
            name_cell.font = font_link
        else:
            name_cell.font = font_normal
        name_cell.fill = fill

        # Get price band information from original data
        price_low_val = row.get('Price Low', '')
        price_high_val = row.get('Price High', '')
        confidence_val = row.get('Confidence', '')

        values = [
            row['Source'],
            row['Type'],
            row['Rarity'],
            row['Attunement'],
            row['Price Formatted'],
            price_low_val,
            price_high_val,
            confidence_val,
            row['Price Source'],
            '⚠️' if row['Is Outlier'] else ''
        ]

        for col_offset, val in enumerate(values, 2):
            cell = ws_main.cell(row=row_idx, column=col_offset, value=val)
            cell.font = font_normal
            cell.fill = fill
            cell.alignment = Alignment(horizontal='left')

    col_widths = [35, 12, 15, 12, 15, 18, 15, 15, 12, 30, 8]
    for col_idx, width in enumerate(col_widths, 1):
        ws_main.column_dimensions[get_column_letter(col_idx)].width = width

    ws_main.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Sheet 2: By Rarity (grouped and sorted)
    ws_rarity = wb.create_sheet(title='By Rarity')
    rarity_headers = ['Rarity', 'Name', 'Source', 'Type', 'Attunement', 'Price (gp)', 'Confidence', 'Notes']
    for col_idx, header in enumerate(rarity_headers, 1):
        cell = ws_rarity.cell(row=1, column=col_idx, value=header)
        cell.fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')

    # Sort by rarity then name
    df_sorted = df.sort_values(['rarity', 'name'])
    for row_idx, (_, row) in enumerate(df_sorted.iterrows(), 2):
        rarity = row['rarity'].replace('_', ' ').title()
        price = row.get('final_price', 0)
        price_formatted = format_price(price) if pd.notna(price) else ''
        confidence = row.get('confidence', '')
        
        values = [
            rarity,
            row['name'],
            row.get('source', ''),
            row.get('type', ''),
            row.get('req_attune', 'none').replace('none', 'No'),
            price_formatted,
            confidence,
            '⚠️' if row.get('is_outlier', False) else ''
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_rarity.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='left')

    # Auto-fit columns for rarity sheet
    for col in ws_rarity.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws_rarity.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 40)

    # Sheet 3: High Confidence (for review)
    ws_high = wb.create_sheet(title='High Confidence')
    high_confidence = df[df.get('confidence', '') == 'High'].sort_values('final_price', ascending=False)
    if len(high_confidence) > 0:
        for col_idx, header in enumerate(headers, 1):
            cell = ws_high.cell(row=1, column=col_idx, value=header)
            cell.fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, (_, row) in enumerate(high_confidence.iterrows(), 2):
            rarity_key = row['rarity'].lower().replace(' ', '_')
            bg_color = RARITY_COLORS.get(rarity_key, 'FFFFFF')
            text_color = RARITY_TEXT_COLORS.get(rarity_key, '000000')
            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
            font_normal = Font(color=text_color, size=10)
            
            price = row.get('final_price', 0)
            
            values = [
                row['name'],
                row.get('source', ''),
                row.get('type', ''),
                row['rarity'].replace('_', ' ').title(),
                row.get('req_attune', 'none').replace('none', 'No'),
                format_price(price) if pd.notna(price) else '',
                row.get('confidence', ''),
                row.get('price_source', '')
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws_high.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_normal
                cell.fill = fill
            # Add URL hyperlink
            if row.get('url'):
                cell = ws_high.cell(row=row_idx, column=1)
                cell.hyperlink = row['url']
                cell.font = Font(color='0563C1', underline='single', size=10)
    else:
        ws_high.cell(row=1, column=1, value="No high confidence items found")

    # Sheet 4: Low Confidence (needs review)
    ws_low = wb.create_sheet(title='Low Confidence')
    low_confidence = df[df.get('confidence', '') == 'Low'].sort_values('final_price', ascending=False)
    if len(low_confidence) > 0:
        for col_idx, header in enumerate(headers, 1):
            cell = ws_low.cell(row=1, column=col_idx, value=header)
            cell.fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, (_, row) in enumerate(low_confidence.iterrows(), 2):
            rarity_key = row['rarity'].lower().replace(' ', '_')
            bg_color = RARITY_COLORS.get(rarity_key, 'FFFFFF')
            text_color = RARITY_TEXT_COLORS.get(rarity_key, '000000')
            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
            font_normal = Font(color=text_color, size=10)
            
            price = row.get('final_price', 0)
            
            values = [
                row['name'],
                row.get('source', ''),
                row.get('type', ''),
                row['rarity'].replace('_', ' ').title(),
                row.get('req_attune', 'none').replace('none', 'No'),
                format_price(price) if pd.notna(price) else '',
                row.get('confidence', ''),
                row.get('price_source', '')
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws_low.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_normal
                cell.fill = fill
            # Add URL hyperlink
            if row.get('url'):
                cell = ws_low.cell(row=row_idx, column=1)
                cell.hyperlink = row['url']
                cell.font = Font(color='0563C1', underline='single', size=10)
    else:
        ws_low.cell(row=1, column=1, value="No low confidence items found")

    wb.save(OUTPUT_XLSX)
    print(f'Saved Excel with 4 sheets to {OUTPUT_XLSX}')

    print(f'\nTotal items: {len(output_rows)}')
    print(f"Hyperlinked items: {sum(1 for r in output_rows if r['URL'])}")

if __name__ == '__main__':
    main()
