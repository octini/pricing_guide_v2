#!/usr/bin/env python3
"""Phase 8: Generate Excel and CSV output"""

import sys
import csv
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

INPUT_CSV = Path('data/processed/items_validated.csv')
OUTPUT_XLSX = Path('output/pricing_guide.xlsx')
OUTPUT_CSV = Path('output/pricing_guide.csv')

# Sourcebook mapping: acronym -> full name (from 5e.tools)
SOURCEBOOK_NAMES = {
    # Core Books
    "PHB": "Player's Handbook",
    "XPHB": "Player's Handbook (2024)",
    "DMG": "Dungeon Master's Guide",
    "XDMG": "Dungeon Master's Guide (2024)",
    "MM": "Monster Manual",
    "XMM": "Monster Manual (2024)",
    
    # Expansion Books
    "XGE": "Xanathar's Guide to Everything",
    "VGM": "Volo's Guide to Monsters",
    "VRGR": "Van Richten's Guide to Ravenloft",
    "TCE": "Tasha's Cauldron of Everything",
    "MTF": "Mordenkainen's Tome of Foes",
    "SCAG": "Sword Coast Adventurer's Guide",
    "GGR": "Guildmasters' Guide to Ravnica",
    "ERLW": "Eberron: Rising from the Last War",
    "EGW": "Explorer's Guide to Wildemount",
    "MOT": "Mythic Odysseys of Theros",
    "FTD": "Fizban's Treasury of Dragons",
    "SCC": "Strixhaven: A Curriculum of Chaos",
    "CRCotN": "Critical Role: Call of the Netherdeep",
    
    # Adventures
    "SKT": "Storm King's Thunder",
    "CoS": "Curse of Strahd",
    "ToA": "Tomb of Annihilation",
    "HotDQ": "Hoard of the Dragon Queen",
    "LMoP": "Lost Mine of Phandelver",
    "WDH": "Waterdeep: Dragon Heist",
    "WDMM": "Waterdeep: Dungeon of the Mad Mage",
    "BGDIA": "Baldur's Gate: Descent Into Avernus",
    "OotA": "Out of the Abyss",
    "PotA": "Princes of the Apocalypse",
    "IDRotF": "Icewind Dale: Rime of the Frostmaiden",
    "GoS": "Ghosts of Saltmarsh",
    "TftYP": "Tales from the Yawning Portal",
    "KftGV": "Keys from the Golden Vault",
    "JttRC": "Journeys through the Radiant Citadel",
    "DSotDQ": "Dragonlance: Shadow of the Dragon Queen",
    "DitLCoT": "Descent into the Lost Caverns of Tsojcanth",
    "LoX": "Light of Xaryxis",
    "PaBTSO": "Phandelver and Below: The Shattered Obelisk",
    "VEoR": "Vecna: Eve of Ruin",
    "WBtW": "The Wild Beyond the Witchlight",
    "QftIS": "Quests from the Infinite Staircase",
    
    # Other Supplements
    "BAM": "Boo's Astral Menagerie",
    "AI": "Acquisitions Incorporated",
    "BGG": "Bigby Presents: Glory of the Giants",
    "BMT": "The Book of Many Things",
    "CM": "Candlekeep Mysteries",
    "AAG": "Astral Adventurer's Guide",
    "SAT": "Sigil and the Outlands",
    "SDW": "Sleeping Dragon's Wake",
    "HotB": "Heroes of the Borderlands",
    "LFL": "Lorwyn: First Light",
    "RoT": "The Rise of Tiamat",
    "RoTOS": "The Rise of Tiamat Online Supplement",
    "RMBRE": "The Lost Dungeon of Rickedness",
    "WttHC": "Stranger Things: Welcome to the Hellfire Club",
    
    # Eberron
    "EET": "Elemental Evil: Trinkets",
    "EFA": "Eberron: Forge of the Artificer",
    "FRAiF": "Forgotten Realms: Adventures in Faerûn",
    "FRHoF": "Forgotten Realms: Heroes of Faerûn",
    "NF": "Netheril's Fall",
    "ExploringEberron24": "Exploring Eberron (2024)",
    "ChroniclesOfEberron": "Chronicles of Eberron",
    
# Third Party / Other
    'DC': 'Divine Contention',
    'FoEQuickstone': 'Frontiers of Eberron: Quickstone',
    'HftT': 'Hunt for the Thessalhydra',
    'MonstersOfDrakkenheim': 'Monsters of Drakkenheim',
    'DungeonsDrakkenheim': 'Dungeons of Drakkenheim',
}


# Item type mapping: 5e.tools code -> common name
TYPE_NAMES = {
    # Standard types
    "M": "Melee Weapon",
    "R": "Ranged Weapon",
    "A": "Ammunition",
    "G": "Adventuring Gear",
    "P": "Potion",
    "S": "Shield",
    "W": "Wondrous Item",
    "OTH": "Other",
    
    # Armor types
    "MA": "Medium Armor",
    "LA": "Light Armor",
    "HA": "Heavy Armor",
    
    # Specific item types
    "SCF": "Spellcasting Focus",
    "AT": "Artisan's Tools",
    "INS": "Musical Instrument",
    "T": "Tool",
    "TG": "Trade Goods",
    "FD": "Food & Drink",
    "GS": "Gaming Set",
    "EXP": "Explosive",
    "MNT": "Mount or Vehicle (Land)",
    
    # Magic item categories
    "RG": "Ring",
    "WD": "Wand",
    "RD": "Rod",
    "SC": "Scroll",
    
    # Vehicles
    "SHP": "Ship/Vehicle (Water)",
    "VEH": "Vehicle (Land)",
    "AIR": "Vehicle (Air)",
    
    # Special
    "SPC": "Species Item",  # From Astral Adventurer's Guide
    "Dele": "Delerium",  # From Monsters of Drakkenheim
    "EM": "Eldritch Machine",  # From Exploring Eberron
    "TAH": "Tack & Harness",
}

RARITY_COLORS = {
    'mundane': 'F5F5F5',
    'common': 'FFFFFF',
    'uncommon': '1EFF00',
    'rare': '0070DD',
    'very_rare': 'A335EE',
    'legendary': 'FF8000',
    'artifact': 'E6CC80',
    'unknown_magic': 'DDDDDD',
    'unknown': 'EEEEEE',
    'varies': 'BBBBBB',
}

RARITY_TEXT_COLORS = {
    'uncommon': '000000',
    'rare': 'FFFFFF',
    'very_rare': 'FFFFFF',
    'legendary': '000000',
    'artifact': '000000',
}

def format_price(price_gp):
    if price_gp < 1:
        return f"{int(price_gp * 100)} cp"
    elif price_gp < 10:
        return f"{price_gp:.1f} gp"
    else:
        return f"{int(price_gp):,} gp"

def determine_price_source_label(row):
    if row.get('price_source') == 'official':
        return 'Official'
    elif row.get('price_confidence') == 'multi':
        return f"Amalgamated ({row.get('price_sources', '')})"
    elif row.get('price_confidence') == 'solo':
        return f"Single source ({row.get('price_sources', '')})"
    else:
        return 'Algorithm'


def translate_source(source_code):
    """Translate sourcebook acronym to full name"""
    if pd.isna(source_code):
        return 'Unknown'
    # Handle pipe-separated sources (multiple sources)
    sources = str(source_code).split('|')
    translated = []
    for s in sources:
        s = s.strip()
        translated.append(SOURCEBOOK_NAMES.get(s, s))
    return ', '.join(translated)


def translate_type(type_code):
    """Translate 5e.tools type code to common name"""
    if pd.isna(type_code):
        return 'Unknown'
    # Type codes can be pipe-separated (e.g., 'M|XPHB')
    types = str(type_code).split('|')
    translated = []
    for t in types:
        t = t.strip()
        # Check if the full code exists first (e.g., 'Dele|MonstersOfDrakkenheim')
        if t in TYPE_NAMES:
            translated.append(TYPE_NAMES[t])
        else:
            # Try just the base code before the pipe
            base = t.split('|')[0]
            translated.append(TYPE_NAMES.get(base, base))
    return ', '.join(translated)

def main():
    df = pd.read_csv(INPUT_CSV)
    print(f'Loaded {len(df)} items')
    
    # Exclude generic variants (items with 'items' field in raw JSON)
    # These are placeholder items like "Horn of Valhalla" that have specific variants
    if 'is_generic_variant' in df.columns:
        before = len(df)
        df = df[~df['is_generic_variant'].fillna(False)]
        print(f'Excluded {before - len(df)} generic variants')

    # Copy prices from alias (original item) to reskin items that duplicate another item.
    # e.g. "Cloak of Shadows" (ExploringEberron24) is an alias of "Cloak of Elvenkind".
    # Reskin items should be priced identically to their original since they have identical mechanics.
    if 'alias' in df.columns:
        name_to_price = dict(zip(df['name'], df['final_price']))
        alias_copies = 0
        for idx, row in df.iterrows():
            alias = row.get('alias', '')
            if not alias or pd.isna(alias) or str(alias).strip() == '':
                continue
            alias_price = name_to_price.get(str(alias).strip())
            if alias_price and pd.notna(alias_price) and alias_price > 0:
                # Always copy the alias price — reskins have identical mechanics to their original
                df.loc[idx, 'final_price'] = alias_price
                df.loc[idx, 'price_low'] = round(alias_price * 0.8, 2)
                df.loc[idx, 'price_high'] = round(alias_price * 1.2, 2)
                alias_copies += 1
        if alias_copies:
            print(f'Copied prices from alias originals to {alias_copies} reskin items')

    output_rows = []
    for _, row in df.iterrows():
        price = row.get('final_price', row.get('rule_price', 0))
        price_low = row.get('price_low', 0)
        price_high = row.get('price_high', 0)
        confidence = row.get('confidence', '')
        # Build notes - flag items without reference sources
        notes = ''
        if row.get('is_outlier', False):
            notes = '⚠️'
        if not row.get('has_reference_source', True):
            notes = notes + ' 🤖' if notes else '🤖'  # Robot emoji for algorithmic items
        
        output_rows.append({
            'Name': row['name'],
            'Source': translate_source(row.get('source', '')),
            'Type Code': row.get('type', ''),  # Keep original for reference
            'Type': translate_type(row.get('type', '')),
            'Rarity': row['rarity'].replace('_', ' ').title(),
            'Attunement': 'Yes' if row.get('req_attune') in ('open', 'class') else 'No',
            'Price (gp)': round(float(price), 2) if pd.notna(price) else 0,
            'Price Formatted': format_price(float(price)) if pd.notna(price) else '0 gp',
            'Price Low': format_price(float(price_low)) if pd.notna(price_low) and price_low > 0 else '',
            'Price High': format_price(float(price_high)) if pd.notna(price_high) and price_high > 0 else '',
            'Confidence': confidence,
            'Price Source': determine_price_source_label(row.to_dict()),
            'URL': row.get('url', ''),
            'Notes': notes,
            'Is Outlier': row.get('is_outlier', False),
            'Has Reference': row.get('has_reference_source', True),
        })

    out_df = pd.DataFrame(output_rows)

    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    # For CSV, include additional columns beyond the main view
    csv_df = out_df.drop(columns=['URL', 'Is Outlier'])
    csv_df.to_csv(OUTPUT_CSV, index=False, quoting=csv.QUOTE_ALL)
    print(f'Saved CSV to {OUTPUT_CSV}')

    # Create multi-sheet Excel workbook
    wb = openpyxl.Workbook()
    
    # Sheet 1: Main Pricing Guide
    ws_main = wb.active
    ws_main.title = 'Pricing Guide'

    headers = ['Name', 'Source', 'Type', 'Rarity', 'Attunement', 'Price (gp)', 'Price Low', 'Price High', 'Confidence', 'Price Source', 'Notes']
    # Note: 'Type Code' is hidden in Excel but available in CSV
    excel_headers = ['Name', 'Source', 'Type', 'Rarity', 'Attunement', 'Price (gp)', 'Price Low', 'Price High', 'Confidence', 'Price Source', 'Notes']
    for col_idx, header in enumerate(excel_headers, 1):
        cell = ws_main.cell(row=1, column=col_idx, value=header)
        cell.fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')

    ws_main.freeze_panes = 'A2'

    for row_idx, row in enumerate(output_rows, 2):
        rarity_key = row['Rarity'].lower().replace(' ', '_')
        bg_color = RARITY_COLORS.get(rarity_key, 'FFFFFF')
        text_color = RARITY_TEXT_COLORS.get(rarity_key, '000000')

        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        font_normal = Font(color=text_color, size=10)
        font_link = Font(color='0563C1', underline='single', size=10)

        name_cell = ws_main.cell(row=row_idx, column=1, value=row['Name'])
        if row['URL']:
            name_cell.hyperlink = row['URL']
            name_cell.font = font_link
        else:
            name_cell.font = font_normal
        name_cell.fill = fill

        # Get price band information from original data
        price_low_val = row.get('Price Low', '')
        price_high_val = row.get('Price High', '')
        confidence_val = row.get('Confidence', '')

        values = [
            row['Source'],
            row['Type'],
            row['Rarity'],
            row['Attunement'],
            row['Price Formatted'],
            price_low_val,
            price_high_val,
            confidence_val,
            row['Price Source'],
            row.get('Notes', '⚠️' if row['Is Outlier'] else '')
        ]

        for col_offset, val in enumerate(values, 2):
            cell = ws_main.cell(row=row_idx, column=col_offset, value=val)
            cell.font = font_normal
            cell.fill = fill
            cell.alignment = Alignment(horizontal='left')

    col_widths = [35, 12, 15, 12, 15, 18, 15, 15, 12, 30, 8]
    for col_idx, width in enumerate(col_widths, 1):
        ws_main.column_dimensions[get_column_letter(col_idx)].width = width

    ws_main.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Sheet 2: By Rarity (grouped and sorted)
    ws_rarity = wb.create_sheet(title='By Rarity')
    rarity_headers = ['Rarity', 'Name', 'Source', 'Type', 'Attunement', 'Price (gp)', 'Confidence', 'Notes']
    for col_idx, header in enumerate(rarity_headers, 1):
        cell = ws_rarity.cell(row=1, column=col_idx, value=header)
        cell.fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')

    # Sort by rarity then name
    df_sorted = df.sort_values(['rarity', 'name'])
    for row_idx, (_, row) in enumerate(df_sorted.iterrows(), 2):
        rarity = row['rarity'].replace('_', ' ').title()
        price = row.get('final_price', 0)
        price_formatted = format_price(price) if pd.notna(price) else ''
        confidence = row.get('confidence', '')
        
        values = [
            rarity,
            row['name'],
            row.get('source', ''),
            row.get('type', ''),
            'Yes' if row.get('req_attune') in ('open', 'class') else 'No',
            price_formatted,
            confidence,
            '⚠️' if row.get('is_outlier', False) else ''
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_rarity.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='left')

    # Auto-fit columns for rarity sheet
    for col in ws_rarity.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws_rarity.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 40)

    # Sheet 3: High Confidence (for review)
    ws_high = wb.create_sheet(title='High Confidence')
    confidence_headers = ['Name', 'Source', 'Type', 'Rarity', 'Attunement', 'Price (gp)', 'Confidence', 'Price Source']
    high_confidence = df[df.get('confidence', '') == 'High'].sort_values('final_price', ascending=False)
    if len(high_confidence) > 0:
        for col_idx, header in enumerate(confidence_headers, 1):
            cell = ws_high.cell(row=1, column=col_idx, value=header)
            cell.fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, (_, row) in enumerate(high_confidence.iterrows(), 2):
            rarity_key = row['rarity'].lower().replace(' ', '_')
            bg_color = RARITY_COLORS.get(rarity_key, 'FFFFFF')
            text_color = RARITY_TEXT_COLORS.get(rarity_key, '000000')
            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
            font_normal = Font(color=text_color, size=10)
            
            price = row.get('final_price', 0)
            
            values = [
                row['name'],
                row.get('source', ''),
                row.get('type', ''),
                row['rarity'].replace('_', ' ').title(),
                row.get('req_attune', 'none').replace('none', 'No'),
                format_price(price) if pd.notna(price) else '',
                row.get('confidence', ''),
                row.get('price_source', '')
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws_high.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_normal
                cell.fill = fill
            # Add URL hyperlink
            if row.get('url'):
                cell = ws_high.cell(row=row_idx, column=1)
                cell.hyperlink = row['url']
                cell.font = Font(color='0563C1', underline='single', size=10)
    else:
        ws_high.cell(row=1, column=1, value="No high confidence items found")

    # Sheet 4: Low Confidence (needs review)
    ws_low = wb.create_sheet(title='Low Confidence')
    low_confidence = df[df.get('confidence', '') == 'Low'].sort_values('final_price', ascending=False)
    if len(low_confidence) > 0:
        for col_idx, header in enumerate(confidence_headers, 1):
            cell = ws_low.cell(row=1, column=col_idx, value=header)
            cell.fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, (_, row) in enumerate(low_confidence.iterrows(), 2):
            rarity_key = row['rarity'].lower().replace(' ', '_')
            bg_color = RARITY_COLORS.get(rarity_key, 'FFFFFF')
            text_color = RARITY_TEXT_COLORS.get(rarity_key, '000000')
            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
            font_normal = Font(color=text_color, size=10)
            
            price = row.get('final_price', 0)
            
            values = [
                row['name'],
                row.get('source', ''),
                row.get('type', ''),
                row['rarity'].replace('_', ' ').title(),
                row.get('req_attune', 'none').replace('none', 'No'),
                format_price(price) if pd.notna(price) else '',
                row.get('confidence', ''),
                row.get('price_source', '')
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws_low.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_normal
                cell.fill = fill
            # Add URL hyperlink
            if row.get('url'):
                cell = ws_low.cell(row=row_idx, column=1)
                cell.hyperlink = row['url']
                cell.font = Font(color='0563C1', underline='single', size=10)
    else:
        ws_low.cell(row=1, column=1, value="No low confidence items found")

    wb.save(OUTPUT_XLSX)
    print(f'Saved Excel with 4 sheets to {OUTPUT_XLSX}')

    print(f'\nTotal items: {len(output_rows)}')
    print(f"Hyperlinked items: {sum(1 for r in output_rows if r['URL'])}")

if __name__ == '__main__':
    main()
